import streamlit as st
import pandas as pd
import plotly.express as px
from io import BytesIO
import openpyxl
import re
from openpyxl.styles import PatternFill

st.set_page_config(layout="wide", initial_sidebar_state="collapsed")

# スタイル設定
st.markdown("""
    <style>
    html, body, [class^="css"]  {
        font-family: Arial, sans-serif !important;
    }
    .stButton button {
        padding: 0.4rem 1rem;
        font-size: 1.1rem;
    }
    </style>
""", unsafe_allow_html=True)

st.title("🖥️ GPUmon & PPM Log Visualizer")

# 2カラムアップローダー
uploader_cols = st.columns(2)
uploaded_gpu_file = uploader_cols[0].file_uploader("GPUmon ログファイル (.txt)", type="txt")
uploaded_ppm_file = uploader_cols[1].file_uploader("PPM ログファイル (.txt)", type="txt")

# Runボタン
if st.button("▶️ Run Conversion"):
    # ===== GPUmon処理 =====
    if uploaded_gpu_file:
        df_gpu = pd.read_csv(uploaded_gpu_file, sep="\t")
        st.success("✅ GPUmonファイルを読み込みました！")
        st.subheader("📄 GPUmon テーブル表示")
        st.dataframe(df_gpu, use_container_width=True)

        def convert_df_to_excel(df):
            output = BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='GPUmon Log')
            return output.getvalue()

        excel_data_gpu = convert_df_to_excel(df_gpu)
        st.download_button("📥 GPUmon Excel出力", data=excel_data_gpu, file_name="GPUmon_Output.xlsx")

        st.subheader("📊 GPUmon Plotlyグラフ")
        col_x = st.selectbox("X軸を選択 (GPUmon)", options=df_gpu.columns, index=0)
        col_y = st.multiselect("Y軸を選択 (GPUmon)", options=df_gpu.columns[1:], default=["Fan1 Current Speed", "Sensor 00"])
        if col_x and col_y:
            fig = px.line(df_gpu, x=col_x, y=col_y)
            st.plotly_chart(fig, use_container_width=True)

    # ===== PPM処理 =====
    if uploaded_ppm_file:
        try:
            content = uploaded_ppm_file.read()
            try:
                text = content.decode("utf-8")
            except UnicodeDecodeError:
                text = content.decode("shift_jis", errors="replace")
            lines = text.splitlines()
        except Exception as e:
            st.error(f"ファイル読み込み中にエラーが発生しました: {e}")
            lines = []

        def get_indent_index(indent):
            return indent // 2 * 2  # Always even index: 0, 2, 4, 6, etc.

        data_rows = []
        for line in lines:
            line = line.rstrip()
            if not line:
                data_rows.append({})
                continue
            indent = len(line) - len(line.lstrip())
            if ":" not in line:
                continue
            key, value = map(str.strip, line.split(":", 1))
            col_key = chr(65 + get_indent_index(indent))  # A, C, E, G
            col_val = chr(65 + get_indent_index(indent) + 1)  # B, D, F, H
            data_rows.append({col_key: key, col_val: value})

        df_ppm = pd.DataFrame(data_rows)
        st.success("✅ PPMファイルを読み込みました！")
        st.subheader("📄 PPM テーブル表示")
        st.dataframe(df_ppm, use_container_width=True)

        output_ppm = BytesIO()
        with pd.ExcelWriter(output_ppm, engine="openpyxl") as writer:
            df_ppm.to_excel(writer, index=False)
            workbook = writer.book
            worksheet = writer.sheets["Sheet1"]
            for col in worksheet.columns:
                worksheet.column_dimensions[col[0].column_letter].width = 47
            fill_gray = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            for row in worksheet.iter_rows():
                if all(cell.value == "" for cell in row):
                    for cell in row:
                        cell.fill = fill_gray

        st.download_button("📥 PPM Excel出力", data=output_ppm.getvalue(), file_name="PPM_Output.xlsx")
