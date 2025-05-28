import sys
import os
import plotly.express as px
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import tempfile
import plotly.io as pio
import xlsxwriter
import pandas as pd
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles.colors import Color
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import Marker
from sensor_correlation_modules.pipeline_module_to_4 import full_logger_ptat_pipeline as pipeline_4
from sensor_correlation_modules.pipeline_module_to_5 import full_logger_ptat_pipeline as pipeline_5

st.set_page_config(layout="wide", initial_sidebar_state="collapsed")

top_col_right = st.columns([8, 1])
with top_col_right[1]:
    st.page_link("main.py", label="🏠 To Main")

# ==== タブ表示・タイトル表示 ====
st.markdown("""
<hr style="
  height: 8px;
  border: none;
  border-radius: 3px;
  background: linear-gradient(to right, #60a760, #9c8eb0, #d69e6a, #dad577, #335f8c, #b10c67, #039CB2);
  margin-top: 10px;
  margin-bottom: 4px;
">
""", unsafe_allow_html=True)

# ==== ✅ タブのフォントサイズを大きくする ====
st.markdown("""
<style>
button[data-baseweb="tab"] > div[data-testid="stMarkdownContainer"] > p {
    font-size: 19px !important;
    font-weight: bold;
    text-align: center;
}
</style>
""", unsafe_allow_html=True)

# st.button用css
st.markdown("""
<style>
button[data-testid="stDownloadButton-template-download"] {
    background-color: #039CB2;
    color: white;
    border: none;
    border-radius: 8px;
    padding: 0.5rem 1rem;
    font-size: 1rem;
    margin-top: 30px;
}
button[data-testid="stDownloadButton-template-download"]:hover {
    background-color: #105d96;
}
</style>
""", unsafe_allow_html=True)
# 横幅3分割で中央カラムにボタン配置
left, center, right = st.columns([2, 3, 2])
with center:
    st.markdown("""
        <style>
        div.stDownloadButton > button {
            width: 100%;
            padding: 1rem;
            font-size: 1.2rem;
            background-color: crimson;
        }
        </style>
    """, unsafe_allow_html=True)

col1, non_col, col2, = st.columns([3,1,2])
with col1:
    st.title("📊 Sensor correlation analyzer")
with col2:
    try:
        template_path = os.path.join(os.path.dirname(__file__), "../sensor_correlation_modules/Result_template.xlsm")

    except NameError:
        template_path = os.path.join(os.getcwd(), "Result_template.xlsm")

    if os.path.exists(template_path):
        with open(template_path, "rb") as f:
            st.download_button(
                label="📥 Download XLSX Template (for Logger-PTAT)",
                data=f.read(),
                file_name="Result_template.xlsm",
                key="template-download"
            )
    else:
        st.warning("❗ Template not found.")
# st.subheader("### ")
col1, col2, non_col,radio_col = st.columns([5, 5, 1, 5])
with col1:    
    st.markdown("### 1️⃣ Upload files for analysis")
    logger_file = st.file_uploader("Logger raw data", type=None)
with col2:
    st.markdown("### ")
    ptat_file = st.file_uploader("pTAT raw data", type=["csv"])
with radio_col:
    st.markdown("### 2️⃣ Select Experiment Split Mode")
    # 初期化フラグを使って、segment selectboxの初期化を制御
    if "segment_defaults_set" not in st.session_state:
        st.session_state.segment_defaults_set = False
    def reset_segment_defaults():
        if st.session_state.split_mode == "4 segments":
            st.session_state.segment_values = ["pTAT+Fur", "pTAT", "Fur", "Prime95", "None"]
        else:
            st.session_state.segment_values = ["pTAT+Fur", "pTAT", "Fur", "Prime95", "pTAT+Fur+Charging"]
        st.session_state.segment_defaults_set = True

    split_mode = st.radio(
        "Choose number of experiment segments",
        ["4 segments", "5 segments"],
        index=0,
        key="split_mode",
        on_change=reset_segment_defaults
    )


st.markdown("### 3️⃣ Select Segment Labels")
segment_labels = ["1st segment", "2nd segment", "3rd segment", "4th segment", "5th segment"]
default_values = ["pTAT+Fur", "pTAT", "Fur", "Prime95", "None"]
select_options = ["None", "pTAT+Fur", "pTAT", "Fur", "Prime95", "pTAT+Fur+Charging"]


# split_modeに応じて、デフォルト値を再初期化
if not st.session_state.segment_defaults_set:
    reset_segment_defaults()

seg_cols = st.columns(5)
selected_segments = []
for i in range(5):
    with seg_cols[i]:
        selected = st.selectbox(
            segment_labels[i],
            options=select_options,
            index=select_options.index(st.session_state.segment_values[i]),
            key=f"segment_select_{i}"
        )
        selected_segments.append(selected)

def add_sensor_correlation_chart_with_markers_only(xlsx_path, col_x, col_y, legend_names, color_map):
    wb = load_workbook(xlsx_path)
    ws = wb["Sensor Correlation"]

    chart = ScatterChart()
    chart.title = "Sensor Correlation Scatter"
    chart.x_axis.title = col_x
    chart.y_axis.title = col_y
    chart.legend.position = 'r'

    max_row = ws.max_row
    max_col = ws.max_column

    for i in range(1, max_col, 2):
        if i + 1 > max_col:
            break

        xvalues = Reference(ws, min_col=i, min_row=2, max_row=max_row)
        yvalues = Reference(ws, min_col=i + 1, min_row=2, max_row=max_row)
        series = Series(yvalues, xvalues, title=legend_names[i // 2])
        
        # マーカー形式でプロット（線なし、色付きマーカー）
        series.marker.symbol = "circle"
        hex_color = color_map.get(legend_names[i // 2], "#000000").replace("#", "")  # default: black
        series.graphicalProperties = GraphicalProperties(solidFill=hex_color)
        chart.series.append(series)

    ws.add_chart(chart, "A8")
    wb.save(xlsx_path)
    return xlsx_path

def add_sensor_correlation_chart_with_colors(excel_path: str, col_x: str, col_y: str, legend_names: list[str], color_map: dict):
    from openpyxl.utils import get_column_letter
    wb = load_workbook(excel_path)
    ws = wb["Sensor Correlation"]

    chart = ScatterChart()
    chart.title = "Sensor Correlation Scatter"
    chart.x_axis.title = col_x
    chart.y_axis.title = col_y
    chart.legend.position = "r"
    chart.style = 18

    max_row = ws.max_row
    max_col = ws.max_column

    for i in range(1, max_col, 2):
        if i + 1 > max_col:
            break

        x_values = Reference(ws, min_col=i, min_row=2, max_row=max_row)
        y_values = Reference(ws, min_col=i+1, min_row=2, max_row=max_row)

        series = Series(y_values, x_values)
        seg_index = i // 2
        if seg_index < len(legend_names):
            name = legend_names[seg_index]
            series.title = SeriesLabel(v=name)

            # ✅ Excelプロットの色を設定（hexコードからカラーに変換）
            hex_color = color_map.get(name, "#000000").replace("#", "")
            hex_color = color_map.get(name, "#000000").replace("#", "")

            # 🔴 マーカーだけにする：線なし、色付きマーカー設定
            series.graphicalProperties = GraphicalProperties()
            series.graphicalProperties.line.noFill = True  # 線を消す
            series.marker = Marker(symbol="circle")
            gp = GraphicalProperties()
            gp.solidFill = hex_color
            gp.line = LineProperties(noFill=True)  # 枠線を消す

            series.marker = Marker(symbol="circle")
            series.marker.graphicalProperties = gp
        chart.series.append(series)

    ws.add_chart(chart, "A8")
    wb.save(excel_path)

def add_sensor_correlation_sheet(excel_path: str, df_corr: pd.DataFrame, col_x: str, col_y: str):
    df_corr = df_corr.dropna(subset=[col_x, col_y])

    if "Experiment" not in df_corr.columns:
        return  # Experiment 컬럼 없으면 처리하지 않음

    segments = df_corr["Experiment"].dropna().unique().tolist()
    experiment_segments = []

    for seg in segments:
        seg_df = df_corr[df_corr["Experiment"] == seg][[col_x, col_y]].reset_index(drop=True).copy()
        seg_df.columns = [f"{col_x} ({seg})", f"{col_y} ({seg})"]
        experiment_segments.append(seg_df)

    if experiment_segments:
        combined_df = pd.concat(experiment_segments, axis=1)

        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            combined_df.to_excel(writer, sheet_name="Sensor Correlation", index=False)
        wb = load_workbook(excel_path)
        if "Sensor Correlation" in wb.sheetnames:
            for sheet in wb.worksheets:
                sheet.sheet_view.tabSelected = False  # 全て非選択に
            wb["Sensor Correlation"].sheet_view.tabSelected = True  # 対象を選択状態に
            wb.active = wb.sheetnames.index("Sensor Correlation")   # アクティブにする
            wb.save(excel_path)

def add_sensor_correlation_chart(excel_path: str, col_x: str, col_y: str, legend_names: list[str]):
    wb = load_workbook(excel_path)
    ws = wb["Sensor Correlation"]

    chart = ScatterChart()
    chart.title = "Sensor Correlation Scatter"
    chart.style = 13
    chart.x_axis.title = col_x
    chart.y_axis.title = col_y
    chart.legend.position = "r"
    chart.graphicalProperties = GraphicalProperties(ln=None)

    max_row = ws.max_row
    max_col = ws.max_column

    for i in range(1, max_col, 2):
        if i + 1 > max_col:
            break

        x_values = Reference(ws, min_col=i, min_row=2, max_row=max_row)
        y_values = Reference(ws, min_col=i+1, min_row=2, max_row=max_row)

        series = Series(y_values, x_values)
        seg_index = i // 2
        if seg_index < len(legend_names):
            series.title = SeriesLabel(v=legend_names[seg_index])
        series.smooth = False
        chart.series.append(series)

    ws.add_chart(chart, "A8")

    wb.save(excel_path)


output_name = ("Merged")

if logger_file and ptat_file:
    # 横に中央配置
    col_left, col_center, col_right = st.columns([2, 3, 2])
    with col_center:
        # CSSで横長スタイルに
        st.markdown("""
            <style>
            div.stButton > button {
                width: 100%;
                padding: 1rem;
                font-size: 1.2rem;
                background-color: #28a745;
                color: white;
                border: none;
                border-radius: 8px;
            }
            div.stButton > button:hover {
                background-color: #218838;
            }
            </style>
        """, unsafe_allow_html=True)

        if st.button("🚀 Run Analysis", key="run-analysis"):
            with tempfile.TemporaryDirectory() as tmpdir:
                logger_path = os.path.join(tmpdir, logger_file.name)
                ptat_path = os.path.join(tmpdir, ptat_file.name)
                output_excel = os.path.join(tmpdir, output_name.strip() + ".xlsx")

                with open(logger_path, "wb") as f:
                    f.write(logger_file.read())
                with open(ptat_path, "wb") as f:
                    f.write(ptat_file.read())

                with st.spinner("Processing..."):
                    full_logger_ptat_pipeline = pipeline_4 if split_mode == "4 segments" else pipeline_5

                    merged_df, _ = full_logger_ptat_pipeline(
                        logger_input_raw=logger_path,
                        ptat_input_raw=ptat_path,
                        merged_excel_output=output_excel
                    )

                if merged_df is not None:
                    st.success("✅ Analysis Complete!")

                    with open(output_excel, "rb") as f:
                        st.session_state["excel_bytes"] = f.read()
                        st.session_state["excel_filename"] = output_name.strip() + ".xlsx"


if "excel_bytes" in st.session_state:

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
        tmp_excel.write(st.session_state["excel_bytes"])
        tmp_excel_path = tmp_excel.name

    try:
        df = pd.read_excel(tmp_excel_path, sheet_name="Experiment Labeled")
        numeric_cols = df.select_dtypes(include='number').columns.tolist()

        tabs = st.tabs(["Skintemp-Sensortemp", "Time-Power"])

        with tabs[0]:
            if len(numeric_cols) >= 2:
                row1_col1, row1_col2 = st.columns(2)

                with row1_col1:
                    col_x = st.selectbox("Select X-axis column", numeric_cols, index=0, key="x1")

                with row1_col2:
                    y_options = [col for col in numeric_cols if col != col_x]
                    current_y = st.session_state.get("y1", y_options[0])
                    if current_y in y_options:
                        y_index = y_options.index(current_y)
                    else:
                        y_index = 0
                    col_y = st.selectbox("Select Y-axis column", y_options, index=y_index, key="y1")

                row2_col1, row2_col2 = st.columns(2)
                with row2_col1:
                    bestp_val = st.number_input("BestP skin spec (deg)", value=0.0)
                with row2_col2:
                    bal_val = st.number_input("Bal skin spec (deg)", value=0.0)

                row3_col1, row3_col2, row3_col3 = st.columns([3,1,4])
                with row3_col1:
                    point_opacity = st.slider("Marker opacity", 0.1, 1.0, 0.7, 0.1, key="opacity1")
                # with row3_col2:
                #     show_y_equals_x = st.checkbox("Show y = x line", value=True)
                with row3_col2:
                    show_grid = st.checkbox("Show grid", value=True)
                with row3_col3:
                    if "Experiment" in df.columns:
                        # セグメント数と選択ラベルを取得
                        num_segments = 4 if split_mode == "4 segments" else 5
                        effective_segments = selected_segments[:num_segments]
                        # Experiment列のユニーク値取得（順序保持）
                        exp_options = df["Experiment"].dropna().unique().tolist()[:num_segments]
                        # ExperimentラベルとUIでの選択ラベルを対応づけ
                        exp_display_map = dict(zip(exp_options, effective_segments))  # 例: {"Exp1": "pTAT+Fur", ...}
                        # UIに表示するラベルリスト（選択肢）
                        exp_display_labels = [exp_display_map[exp] for exp in exp_options]
                        # ユーザーに表示されるselectは変換名（選択ラベル）、選択された値を元のExperimentに戻す
                        selected_display_labels = st.multiselect("Filter by Experiment", exp_display_labels, default=exp_display_labels)
                        selected_exps = [exp for exp, label in exp_display_map.items() if label in selected_display_labels]
                        # dfをフィルター
                        df_filtered = df[df["Experiment"].isin(selected_exps)]
                    else:
                        df_filtered = df.copy()
                # 色をPlotlyの順番で固定
                plotly_colors = px.colors.qualitative.Plotly
                color_map = {seg: plotly_colors[i % len(plotly_colors)] for i, seg in enumerate(effective_segments)}
                effective_segments = selected_segments[:num_segments]
    
                # color_map = {
                #     1st: "#1f77b4",
                #     2nd: "#2ca02c",
                #     3rd: "#ff7f0e",
                #     4th: "#d62728",
                #     5th: "#9467bd"
                # }

                num_segments = 4 if split_mode == "4 segments" else 5
                fixed_colors = ["#1f77b4", "#2ca02c", "#ff7f0e", "#d62728", "#9467bd"]
                num_segments = 4 if split_mode == "4 segments" else 5
                effective_segments = selected_segments[:num_segments]
                color_map = {label: fixed_colors[i] for i, label in enumerate(effective_segments)}
                fig = go.Figure()
                
                fig.add_vline(
                    x=bal_val,
                    line=dict(color="green", dash="dot", width=3),
                    annotation_text="Bal spec",
                    annotation_position="top left",
                    layer="above",
                    opacity=0.7
                )

                fig.add_vline(
                    x=bestp_val,
                    line=dict(color="red", dash="dot", width=3),
                    annotation_text="BestP spec",
                    annotation_position="top right",
                    layer="above",
                    opacity=0.7
                )

                if "Experiment" in df.columns:
                    unique_exps = df_filtered["Experiment"].dropna().unique().tolist()
                    px_colors = px.colors.qualitative.Plotly
                    
                    for i, exp in enumerate(unique_exps):
                        if i >= len(effective_segments):
                            break
                        seg_label = effective_segments[i]
                        exp_df = df_filtered[df_filtered["Experiment"] == exp]
                        fig.add_trace(go.Scatter(
                            x=exp_df[col_x],
                            y=exp_df[col_y],
                            mode="markers",
                            name=seg_label,
                            marker=dict(color=color_map[seg_label]),
                            opacity=point_opacity
                        ))

                fig.add_vline(
                    x=bal_val,
                    line=dict(color="green", dash="dot"),
                    annotation_text="Bal spec",
                    annotation_position="top left"
                )

                fig.add_vline(
                    x=bestp_val,
                    line=dict(color="red", dash="dot"),
                    annotation_text="BestP spec",
                    annotation_position="top right"
                )

                # if show_y_equals_x:
                fig.add_trace(go.Scatter(
                    x=[25, 60],
                    y=[25, 60],
                    mode="lines",
                    line=dict(color="cyan"),
                    name="y = x"
                ))

                fig.update_layout(
                    xaxis=dict(title=col_x, range=[25, 60], title_font=dict(size=18), tickfont=dict(size=14)),
                    yaxis=dict(title=col_y, range=[25, 60], title_font=dict(size=18), tickfont=dict(size=14)),
                    legend=dict(title="Experiment", font=dict(size=14)),
                    width=800,
                    height=870,
                    template="simple_white",
                    title=""
                )
                fig.update_xaxes(showgrid=show_grid)
                fig.update_yaxes(showgrid=show_grid)

                with st.spinner("Processing Sensor Correlation Chart and Output..."):
                    # Sensor Correlation シートを追加
                    add_sensor_correlation_sheet(
                        excel_path=tmp_excel_path,
                        df_corr=df_filtered,
                        col_x=col_x,
                        col_y=col_y
                    )

                    # Sensor Correlation グラフをA8に追加
                    add_sensor_correlation_chart(
                        excel_path=tmp_excel_path,
                        col_x=col_x,
                        col_y=col_y,
                        legend_names=effective_segments
                    )

                    add_sensor_correlation_chart_with_colors(
                        excel_path=tmp_excel_path,
                        col_x=col_x,
                        col_y=col_y,
                        legend_names=effective_segments,  # 例: ["pTAT+Fur", "pTAT", ...]
                        color_map=color_map               # Plotlyで使ったのと同じ辞書
                    )

                    # 💾 ボタン（グラフの直前に配置）
                    with open(tmp_excel_path, "rb") as f_corr:
                        st.download_button(
                            label="📥 Output XLSX Sensor Correlation Sheet",
                            data=f_corr.read(),
                            file_name="Merged_with_Correlation.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="centered-download"
                        )
                    st.plotly_chart(fig, use_container_width=True)

                    # try:
                    #     png_bytes = pio.to_image(fig, format="png")
                    #     st.download_button("📥 Download Chart as PNG", data=png_bytes, file_name="chart.png", mime="image/png")
                    # except Exception:
                    #     st.warning("⚠️ Install 'kaleido' for PNG export.")

                    with st.expander("📋 View Raw Data", expanded=True):
                        cols_to_show = [col_x, col_y]
                        if "Experiment" in df.columns:
                            cols_to_show.append("Experiment")
                        st.dataframe(df[cols_to_show])
           

        with tabs[1]:
            power_cols = [col for col in df.columns if any(key in col for key in ["IA", "GT", "Package"])]
            time_col = next((col for col in df.columns if "Time" in col), None)

            if power_cols and time_col:
                fig2 = go.Figure()
                for col in power_cols:
                    fig2.add_trace(go.Scatter(x=df[time_col], y=df[col], mode="lines", name=col))
                fig2.update_layout(
                    xaxis=dict(title=time_col, title_font=dict(size=18), tickfont=dict(size=14)),
                    yaxis=dict(title="Power (W)", title_font=dict(size=18), tickfont=dict(size=14)),
                    legend=dict(font=dict(size=14)),
                    width=900,
                    height=600,
                    template="simple_white",
                    title=""
                )
                st.plotly_chart(fig2, use_container_width=True)

    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
